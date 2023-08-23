import openpyxl

# Carrega a planinha
planinha_toda = openpyxl.load_workbook("automacao_excel_1.xlsx")

# Pagina da planinha a ser editada
planinha = planinha_toda.active

# cria uma array pra guardar as informações(nomes)
lista_empregados = []

# fudendo no talo no cu da safada
for row_number, linha in enumerate(planinha.iter_rows(min_row=2, max_row=planinha.max_row, values_only=True), start=2):
    soma = linha[1] + linha[2]
    if soma != linha[3]:
        lista_empregados.append(linha[0])
        # Update the cell in column 5 (E) with the calculated value
        planinha.cell(row=row_number, column=5, value=soma)
    else:
        planinha.cell(row=row_number, column=5, value="Ok nigga")

        
planinha_toda.save('teste_1.xlsx')

planinha_final = openpyxl.Workbook()

funcionario = planinha_final['Sheet']

for row, info in enumerate(lista_empregados, start=1):
    funcionario['A'+str(row)] = "Empregado burro:"
    funcionario['B'+str(row)] = info
    


planinha_final.save('Funcionarios invalidos.xlsx')

